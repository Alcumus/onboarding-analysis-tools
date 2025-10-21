import argparse
import csv
import re
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from fuzzywuzzy import fuzz
from datetime import datetime, timedelta
from convertTimeZone import convertFromIANATimezone

CBX_DEFAULT_STANDARD_SUBSCRIPTION = 803
CBX_HEADER_LENGTH = 28
# noinspection SpellCheckingInspection
CBX_ID, CBX_COMPANY_FR, CBX_COMPANY_EN, CBX_COMPANY_OLD, CBX_ADDRESS, CBX_CITY, CBX_STATE, \
    CBX_COUNTRY, CBX_ZIP, CBX_FISTNAME, CBX_LASTNAME, CBX_EMAIL, CBX_EXPIRATION_DATE, CBX_REGISTRATION_STATUS, \
    CBX_SUSPENDED, CBX_MODULES, CBX_ACCESS_MODES, CBX_ACCOUNT_TYPE, CBX_SUB_PRICE_CAD, CBX_EMPL_PRICE_CAD,\
    CBX_SUB_PRICE_USD, CBX_EMPL_PRICE_USD, CBX_HIRING_CLIENT_NAMES, \
    CBX_HIRING_CLIENT_IDS, CBX_HIRING_CLIENT_QSTATUS, CBX_PARENTS, CBX_ASSESSMENT_LEVEL, CBX_IS_NEW_PRODUCT = range(CBX_HEADER_LENGTH)

HC_HEADER_LENGTH = 41
HC_COMPANY, HC_FIRSTNAME, HC_LASTNAME, HC_EMAIL, HC_CONTACT_PHONE, HC_CONTACT_LANGUAGE, HC_STREET, HC_CITY, \
    HC_STATE, HC_COUNTRY, HC_ZIP, HC_CATEGORY, HC_DESCRIPTION, HC_PHONE, HC_EXTENSION, HC_FAX,  HC_WEBSITE,\
    HC_LANGUAGE, HC_IS_TAKE_OVER, HC_TAKEOVER_QUALIFICATION_DATE, HC_TAKEOVER_QF_STATUS, \
    HC_PROJECT_NAME, HC_QUESTIONNAIRE_NAME, HC_QUESTIONNAIRE_ID, HC_PRICING_GROUP_ID, HC_PRICING_GROUP_CODE, \
    HC_HIRING_CLIENT_NAME, HC_HIRING_CLIENT_ID, HC_IS_ASSOCIATION_FEE, HC_BASE_SUBSCRIPTION_FEE, \
    HC_CONTACT_CURRENCY, HC_AGENT_IN_CHARGE_ID, HC_TAKEOVER_FOLLOW_UP_DATE, HC_TAKEOVER_RENEWAL_DATE, \
    HC_INFORMATION_SHARED, HC_CONTACT_TIMEZONE, HC_DO_NOT_MATCH, HC_FORCE_CBX_ID, HC_AMBIGUOUS, \
    HC_CONTRACTORCHECK_ACCOUNT, HC_ASSESSMENT_LEVEL \
    = range(HC_HEADER_LENGTH)

SUPPORTED_CURRENCIES = ('CAD', 'USD')

def normalize_postal_code(code):
    if not code:
        return ''
    import unicodedata
    code = str(code).strip().upper()
    code = unicodedata.normalize('NFKD', code)
    code = ''.join([c for c in code if not unicodedata.combining(c)])
    code = re.sub(r'[^A-Z0-9]', '', code)
    return code

def calculate_location_bonus(input_address, input_city, input_province, candidate_address, candidate_city, candidate_province, input_country=None, candidate_country=None):
    """Calculate location proximity bonus for business scoring."""
    if not input_address or not candidate_address:
        return 0
        
    # Exact address match (same building/suite)
    if normalize_address(input_address) == normalize_address(candidate_address):
        return 25  # Increased from 20
    
    # Same city match
    if (input_city and candidate_city and 
        input_city.lower().strip() == candidate_city.lower().strip()):
        return 15  # Increased from 10
        
    # Same province/state match
    if (input_province and candidate_province and 
        input_province.lower().strip() == candidate_province.lower().strip()):
        return 10  # Increased from 5
        
    # Same country match (important for international vs domestic)
    if (input_country and candidate_country and 
        input_country.lower().strip() == candidate_country.lower().strip()):
        return 8   # Country match bonus
        
    # PENALTY: International mismatch (domestic contractor vs international candidate)
    # This helps prevent Canadian contractors from being matched to Swiss/other international companies
    if (input_country and candidate_country and 
        input_country.lower().strip() != candidate_country.lower().strip()):
        return -15  # Heavy penalty for country mismatch
        
    return 0

def normalize_address(address):
    """Normalize address for exact matching comparison."""
    if not address:
        return ''
    # Normalize unicode, remove extra spaces, standardize punctuation
    import unicodedata
    normalized = unicodedata.normalize('NFKD', str(address).strip().lower())
    normalized = ''.join([c for c in normalized if not unicodedata.combining(c)])
    # Standardize common address abbreviations and punctuation
    normalized = re.sub(r'\bboul\.?\b', 'boulevard', normalized)
    normalized = re.sub(r'\brue\b', 'street', normalized) 
    normalized = re.sub(r'\bste\.?\b', 'suite', normalized)
    normalized = re.sub(r'\bbur\.?\b', 'bureau', normalized)
    normalized = re.sub(r'[^\w\s]', ' ', normalized)  # Remove punctuation
    normalized = re.sub(r'\s+', ' ', normalized).strip()  # Normalize spaces
    return normalized

assessment_levels = {
    "gold": 2,
    "silver": 2,
    "bronze" : 1,
    "level3": 2, 
    "level2": 2,
    "level1": 1,
    "3":2,
    "2":2,
    "1":1
}

# Used in order to switch code and id in data to import
rd_pricing_group_id_col = -1
rd_pricing_group_code_col = -1

# noinspection SpellCheckingInspection
cbx_headers = ['id', 'name_fr', 'name_en', 'old_names', 'address', 'city', 'state', 'country', 'postal_code',
               'first_name', 'last_name', 'email', 'cbx_expiration_date', 'registration_code', 'suspended',
               'modules', 'access_modes', 'code', 'subscription_price_cad', 'employee_price_cad',
               'subscription_price_usd', 'employee_price_usd', 'hiring_client_names',
               'hiring_client_ids', 'hiring_client_qstatus', 'parents', 'assessment_level', 'new_product']

# noinspection SpellCheckingInspection
hiring_client_headers = ['contractor_name', 'contact_first_name', 'contact_last_name', 'contact_email', 'contact_phone',
              'contact_language', 'address', 'city', 'province_state_iso2', 'country_iso2',
              'postal_code', 'category', 'description', 'phone', 'extension', 'fax', 'website', 'language',
              'is_take_over', 'qualification_expiration_date',
              'qualification_status', 'batch', 'questionnaire_name', 'questionnaire_id',
              'pricing_group_id', 'pricing_group_code', 'hiring_client_name', 'hiring_client_id', 'is_association_fee',
              'base_subscription_fee', 'contact_currency', 'agent_in_charge_id', 'take_over_follow-up_date',
              'renewal_date', 'information_shared', 'contact_timezone', 'do_not_match',
              'force_cbx_id', 'ambiguous', 'contractorcheck_account', 'assessment_level']

# noinspection SpellCheckingInspection
analysis_headers = ['cbx_id', 'hc_contractor_summary', 'analysis','cbx_contractor', 'cbx_street', 'cbx_city', 'cbx_state', 'cbx_zip', 'cbx_country',
                    'cbx_expiration_date', 'registration_status', 'suspended', 'cbx_email',
                    'cbx_first_name', 'cbx_last_name', 'modules', 'cbx_account_type',
                    'cbx_subscription_fee', 'cbx_employee_price', 'parents', 'previous',
                    'hiring_client_names', 'hiring_client_count',
                    'is_in_relationship', 'is_qualified', 'ratio_company', 'ratio_address',
                    'contact_match', 'cbx_assessment_level', 'new_product', 'generic_domain', 'match_count', 'match_count_with_hc',
                    'is_subscription_upgrade', 'upgrade_price', 'prorated_upgrade_price', 'create_in_cbx',
                    'action', 'index']

rd_headers = ['contractor_name', 'contact_first_name', 'contact_last_name', 'contact_email', 'contact_phone',
              'contact_language', 'address', 'city', 'province_state_iso2', 'country_iso2',
              'postal_code', 'description', 'phone', 'extension', 'fax', 'website', 'language',
              'qualification_expiration_date', 'qualification_status', 'contact_currency',
              'agent_in_charge_id', 'renewal_date', 'information_shared', 'contact_timezone', 'questionnaire_name', 'questionnaire_ids',
              'pricing_group_code', 'pricing_group_id', 'hiring_client_id', 'contractorcheck_account', 'assessment_level']

existing_contractors_headers = ['cbx_id']
existing_contractors_headers.extend(rd_headers.copy())

hubspot_headers = ['contractor_name', 'contact_first_name', 'contact_last_name', 'contact_email', 'contact_phone',
              'contact_language', 'address', 'city', 'province_state_iso2', 'country_iso2',
              'postal_code', 'cbx_id', 'cbx_expiration_date', 'questionnaire_name',
              'questionnaire_id', 'hiring_client_name', 'hiring_client_id', 'action']


metadata_headers = ['metadata_x', 'metadata_y', 'metadata_z', '...']

# noinspection SpellCheckingInspection
BASE_GENERIC_DOMAIN = ['yahoo.ca', 'yahoo.com', 'hotmail.com', 'gmail.com', 'outlook.com',
                       'bell.com', 'bell.ca', 'videotron.ca', 'eastlink.ca', 'kos.net', 'bellnet.ca', 'sasktel.net',
                       'aol.com', 'tlb.sympatico.ca', 'sogetel.net', 'cgocable.ca',
                       'hotmail.ca', 'live.ca', 'icloud.com', 'hotmail.fr', 'yahoo.com', 'outlook.fr', 'msn.com',
                       'globetrotter.net', 'live.com', 'sympatico.ca', 'live.fr', 'yahoo.fr', 'telus.net',
                       'shaw.ca', 'me.com', 'bell.net', 'cablevision.qc.ca', 'live.ca', 'tlb.sympatico.ca',
                       '', 'videotron.qc.ca', 'ivic.qc.ca', 'qc.aira.com', 'canada.ca', 'axion.ca', 'bellsouth.net', 
                       'telusplanet.net','rogers.com', 'mymts.net', 'nb.aibn.com', 'on.aibn.com', 'live.be', 'nbnet.nb.ca',
                       'execulink.com', 'bellaliant.com', 'nf.aibn.com', 'clintar.com', 'pathcom.com', 'oricom.ca', 'mts.net',
                       'xplornet.com', 'mcsnet.ca', 'att.net', 'ymail.com', 'mail.com', 'bellaliant.net', 'ns.sympatico.ca', 
                       'ns.aliantzinc.ca', 'mnsi.net']
# noinspection SpellCheckingInspection
BASE_GENERIC_COMPANY_NAME_WORDS = ['construction', 'contracting', 'industriel', 'industriels', 'service',
                                   'services', 'inc', 'limited', 'ltd', 'ltee', 'ltée', 'co', 'industrial',
                                   'solutions', 'llc', 'enterprises', 'systems', 'industries',
                                   'technologies', 'company', 'corporation', 'installations', 'enr', 'inc']


def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


hiring_client_headers_with_metadata = hiring_client_headers.copy()
hiring_client_headers_with_metadata.extend(metadata_headers)
cbx_headers_text = '\n'.join([', '.join(x) for x in list(chunks(cbx_headers, 5))])
hiring_client_headers_text = '\n'.join([', '.join(x) for x in list(chunks(hiring_client_headers_with_metadata, 5))])
analysis_headers_text = '\n'.join([', '.join(x) for x in list(chunks(analysis_headers, 5))])
existing_contractors_text = '\n'.join([', '.join(x) for x in list(chunks(existing_contractors_headers, 5))])

if len(hiring_client_headers) != HC_HEADER_LENGTH:
    raise AssertionError('hc header inconsistencies')

if len(cbx_headers) != CBX_HEADER_LENGTH:
    raise AssertionError('cbx header inconsistencies')

# define commandline parser
parser = argparse.ArgumentParser(
    description='Tool to match contractor list provided by hiring clients to business units in CBX, '
                'all input/output files must be in the current directory',
    formatter_class=argparse.RawTextHelpFormatter)
parser.add_argument('cbx_list',
                    help=f'csv DB export file of business units with the following columns:\n{cbx_headers_text}\n\n')

parser.add_argument('hc_list',
                    help=f'xlsx file of the hiring client contractors and the '
                         f'following columns:\n{hiring_client_headers_text}\n\n')
parser.add_argument('output',
                    help=f'the xlsx file to be created with the hc_list columns and the following analysis columns:'
                         f'\n{analysis_headers_text}\n\n**Please note that metadata columns from the'
                         f' hc file are moved after the analysis data')


parser.add_argument('--hc_list_sheet_name', dest='hc_list_sheet_name', action='store',
                    default=None,
                    help='specify the sheet in the excel file where the hiring client data is located'
                         ' (default separator is the active sheet)')

parser.add_argument('--hc_list_offset', dest='hc_list_offset', action='store',
                    default=None,
                    help='offset where the data starts in the form of <row>,<column> (default is 1,1). '
                         'This includes the headers')

parser.add_argument('--min_company_match_ratio', dest='ratio_company', action='store',
                    default=80,
                    help='Minimum match ratio for contractors, between 0 and 100 (default 80)')

parser.add_argument('--min_address_match_ratio', dest='ratio_address', action='store',
                    default=80,
                    help='Minimum match ratio for addresses (street + zip), between 0 and 100 (default 80)')

parser.add_argument('--additional_generic_domain', dest='additional_generic_domain', action='store',
                    default='',
                    help='list of domains to ignore separated by the list separator (default separator is ;)')

parser.add_argument('--additional_generic_name_word', dest='additional_generic_name_word', action='store',
                    default='',
                    help='list of generic words in company name to ignore separated by the list separator'
                         ' (default separator is ;)')

parser.add_argument('--cbx_list_encoding', dest='cbx_encoding', action='store',
                    default='utf-8-sig',
                    help='Encoding for the cbx list (default: utf-8-sig)')

parser.add_argument('--list_separator', dest='list_separator', action='store',
                    default=';',
                    help='string separator used for lists (default: ;)')

parser.add_argument('--no_headers', dest='no_headers', action='store_true',
                    help='to indicate that input files have no headers')

parser.add_argument('--ignore_warnings', dest='ignore_warnings', action='store_true',
                    help='to ignore data consistency checks and run anyway...')

args = parser.parse_args()
GENERIC_DOMAIN = BASE_GENERIC_DOMAIN + args.additional_generic_domain.split(args.list_separator)
GENERIC_COMPANY_NAME_WORDS = BASE_GENERIC_COMPANY_NAME_WORDS + \
                             args.additional_generic_name_word.split(args.list_separator)


def smart_boolean(bool_data):
    if isinstance(bool_data, str):
        bool_data = bool_data.lower().strip()
        return True if bool_data in ('true', '=true', 'yes', 'vraie', '=vraie', '1') else False
    else:
        return bool(bool_data)


# noinspection PyShadowingNames

def add_analysis_data(hc_row, cbx_row, analysis_string='', ratio_company=0, ratio_address=0, contact_match=False):
    cbx_company = cbx_row[CBX_COMPANY_FR] if cbx_row[CBX_COMPANY_FR] else cbx_row[CBX_COMPANY_EN]
    hc_email = hc_row[HC_EMAIL] if HC_EMAIL < len(hc_row) else ''
    print('   --> ', cbx_company, hc_email, cbx_row[CBX_ID], ratio_company, ratio_address, contact_match)
    hiring_clients_list = cbx_row[CBX_HIRING_CLIENT_NAMES].split(args.list_separator)
    hiring_clients_qstatus = cbx_row[CBX_HIRING_CLIENT_QSTATUS].split(args.list_separator)
    hc_count = len(hiring_clients_list) if cbx_row[CBX_HIRING_CLIENT_NAMES] else 0
    is_in_relationship = True if (
            hc_row[HC_HIRING_CLIENT_NAME] in hiring_clients_list and hc_row[HC_HIRING_CLIENT_NAME]) else False
    is_qualified = False
    sub_price_usd = float(cbx_row[CBX_SUB_PRICE_USD]) if cbx_row[CBX_SUB_PRICE_USD] else 0.0
    employee_price_usd = float(cbx_row[CBX_EMPL_PRICE_USD]) if cbx_row[CBX_EMPL_PRICE_USD] else 0.0
    sub_price_cad = float(cbx_row[CBX_SUB_PRICE_CAD]) if cbx_row[CBX_SUB_PRICE_CAD] else 0.0
    employee_price_cad = float(cbx_row[CBX_EMPL_PRICE_CAD]) if cbx_row[CBX_EMPL_PRICE_CAD] else 0.0
    hiring_client_contractor_summary = f'{hc_row[HC_COMPANY]}, {hc_row[HC_STREET]}, {hc_row[HC_CITY]}, {hc_row[HC_STATE]}, {hc_row[HC_COUNTRY]}, {hc_row[HC_ZIP]}, {hc_row[HC_EMAIL]}, {hc_row[HC_FIRSTNAME]} {hc_row[HC_LASTNAME]}'

    if hc_row[HC_CONTACT_CURRENCY] != '' and hc_row[HC_CONTACT_CURRENCY] not in SUPPORTED_CURRENCIES:
        raise AssertionError(f'Invalid currency: {hc_row[HC_CONTACT_CURRENCY]}, must be in {SUPPORTED_CURRENCIES}')
    for idx, val in enumerate(hiring_clients_list):
        if val == hc_row[HC_HIRING_CLIENT_NAME] and hiring_clients_qstatus[idx] == 'validated':
            is_qualified = True
            break
    try:
        expiration_date = datetime.strptime(cbx_row[CBX_EXPIRATION_DATE],
                                        "%d/%m/%y") if cbx_row[CBX_EXPIRATION_DATE] else None
    except ValueError:
        expiration_date = datetime.strptime(cbx_row[CBX_EXPIRATION_DATE],
                                        "%d/%m/%Y") if cbx_row[CBX_EXPIRATION_DATE] else None

    return {'cbx_id': int(cbx_row[CBX_ID]), 'hc_contractor_summary': hiring_client_contractor_summary, 'analysis':'', 'company': cbx_company, 'address': cbx_row[CBX_ADDRESS],
            'city': cbx_row[CBX_CITY], 'state': cbx_row[CBX_STATE], 'zip': cbx_row[CBX_ZIP],
            'country': cbx_row[CBX_COUNTRY], 'expiration_date': expiration_date,
            'registration_status': cbx_row[CBX_REGISTRATION_STATUS],
            'suspended': cbx_row[CBX_SUSPENDED], 'email': cbx_row[CBX_EMAIL], 'first_name': cbx_row[CBX_FISTNAME],
            'last_name': cbx_row[CBX_LASTNAME], 'modules': cbx_row[CBX_MODULES],
            'account_type': cbx_row[CBX_ACCOUNT_TYPE],
            'subscription_price': sub_price_cad if hc_row[HC_CONTACT_CURRENCY] == "CAD" else sub_price_usd,
            'employee_price': employee_price_cad if hc_row[HC_CONTACT_CURRENCY] == "CAD" else employee_price_usd,
            'parents': cbx_row[CBX_PARENTS], 'previous': cbx_row[CBX_COMPANY_OLD],
            'hiring_client_names': cbx_row[CBX_HIRING_CLIENT_NAMES], 'hiring_client_count': hc_count,
            'is_in_relationship': is_in_relationship, 'is_qualified': is_qualified,
            'ratio_company': ratio_company, 'ratio_address': ratio_address, 'contact_match': contact_match, 
            'cbx_assessment_level': cbx_row[CBX_ASSESSMENT_LEVEL],
            'new_product': cbx_row[CBX_IS_NEW_PRODUCT]
            }


def core_mandatory_provided(hcd):
    mandatory_fields = (HC_COMPANY, HC_FIRSTNAME, HC_LASTNAME, HC_EMAIL, HC_CONTACT_PHONE,
                        HC_STREET, HC_CITY, HC_STATE, HC_COUNTRY, HC_ZIP)
    country = hcd[HC_COUNTRY].strip().lower() if isinstance(hcd[HC_COUNTRY], str) else hcd[HC_COUNTRY]
    for field in mandatory_fields:
        f_value = hcd[field].strip() if isinstance(hcd[field], str) else hcd[field]
        if f_value == "":
            if field == HC_STATE and country not in ('ca', 'us'):
                pass
            else:
                return False
    return True


# noinspection PyShadowingNames
def action(hc_data, cbx_data, create, subscription_update, expiration_date, is_qualified, ignore):
    # FINAL CORRECTED BUSINESS LOGIC: missing_info check comes FIRST
    # Rule: If HC data is incomplete, return missing_info regardless of CBX match status
    # Only proceed with business logic if HC data is complete enough for processing
    
    # Check if this is a new contractor creation case
    if create and not core_mandatory_provided(hc_data):
        return 'missing_info'
    
    # If HC data is complete (or this is an existing contractor), proceed with CBX-based business logic
    
    # Only proceed with business logic if HC data is complete
    has_cbx_match = cbx_data is not None
    
    # Check for association fee first, regardless of create status
    if hc_data[HC_IS_ASSOCIATION_FEE] and cbx_data:
        reg_status = cbx_data.get('registration_status')
        if reg_status in ('Active', 'Non Member') and not cbx_data.get('is_in_relationship', False):
            if expiration_date:
                in_60_days = datetime.now() + timedelta(days=60)
                if expiration_date > in_60_days:
                    return 'association_fee'
                else:
                    return 'add_questionnaire'
            else:
                return 'association_fee'

    if create:
        if smart_boolean(hc_data[HC_IS_TAKE_OVER]):
            return 'activation_link'
        else:
            if hc_data[HC_AMBIGUOUS]:
                return 'ambiguous_onboarding'
            elif core_mandatory_provided(hc_data):
                return 'onboarding'
            else:
                # Should not reach here due to earlier missing_info check, but keep as fallback
                return 'missing_info'
    else:
        reg_status = cbx_data['registration_status']
        if smart_boolean(hc_data[HC_IS_TAKE_OVER]):
            if reg_status == 'Suspended':
                return 'restore_suspended'
            elif reg_status == 'Active':
                return 'add_questionnaire'
            elif reg_status == 'Non Member':
                return 'activation_link'
            else:
                print(f'WARNING: invalid registration status {hc_data[CBX_REGISTRATION_STATUS]}')
                if not ignore:
                    exit(-1)
        else:
            if reg_status == 'Active':
                if cbx_data['is_in_relationship']:
                    if is_qualified:
                        return 'already_qualified'
                    else:
                        return 'follow_up_qualification'
                else:
                    if subscription_update:
                        return 'subscription_upgrade'
                    elif hc_data[HC_IS_ASSOCIATION_FEE] and not cbx_data['is_in_relationship']:
                        if expiration_date:
                            in_60_days = datetime.now() + timedelta(days=60)
                            if expiration_date > in_60_days:
                                return 'association_fee'
                            else:
                                return 'add_questionnaire'
                        else:
                            return 'association_fee'
                    else:
                        return 'add_questionnaire'
            elif reg_status == 'Suspended':
                return 'restore_suspended'
            elif reg_status in ('Non Member', '', None):
                return 're_onboarding'
            else:
                raise AssertionError(f'invalid registration status: {reg_status}')


def remove_generics(company_name):
    for word in GENERIC_COMPANY_NAME_WORDS:
        company_name = re.sub(r'\b' + word + r'\b', '', company_name)
    return company_name


# noinspection PyShadowingNames
def check_headers(headers, standards, ignore):
    headers = [x.lower().strip() for x in headers]
    for idx, val in enumerate(standards):
        if val != headers[idx]:
            print(f'WARNING: got "{headers[idx]}" while expecting "{val}" in column {idx + 1}')
            if not ignore:
                exit(-1)


def clean_company_name(name):
    name = name.lower().replace('.', '').replace(',', '').strip()
    name = re.sub(r"\([^()]*\)", "", name)
    name = remove_generics(name)
    name = re.sub(r'\s+', ' ', name)
    # Remove generic legal suffixes AND common filler words that don't help matching
    words = [w for w in name.split() if w not in ('inc', 'ltd', 'ltée', 'ltee', 'co', 'corp', 'corporation', 'company', 'llc', 'sarl', 'sa', 'plc', 'enr', 'industriel', 'industriels', 'services', 'service', 'solutions', 'systems', 'technologies', 'installations')]
    # MATCHING FIX: Preserve word order instead of sorting alphabetically - word order is crucial for fuzzy matching
    # Remove duplicates while preserving order
    seen = set()
    words_ordered = []
    for word in words:
        if word not in seen:
            seen.add(word)
            words_ordered.append(word)
    return ' '.join(words_ordered).strip()
    if not clean_name:
        return cbx_data
    
    # Extract first few characters for quick filtering
    name_start = clean_name[:3] if len(clean_name) >= 3 else clean_name
    name_length = len(clean_name)
    
    # Common French/English company prefixes that should be ignored for first-char matching
    common_prefixes = ['les ', 'le ', 'la ', 'l ', 'entreprises ', 'entreprise ', 'ets ', 'portes ', 'service ', 'services ', 'groupe ']
    
    # Extract meaningful words (for acronym/word matching)
    hc_words = set(clean_name.split())
    
    # Check if HC name might be an acronym (short, all caps-like, no common words)
    is_likely_acronym = name_length <= 6 and len(hc_words) == 1
    
    # Fast pre-filtering criteria
    filtered = []
    for cbx_row in cbx_data:
        cbx_en = clean_company_name(cbx_row[CBX_COMPANY_EN])
        cbx_fr = clean_company_name(cbx_row[CBX_COMPANY_FR])
        cbx_old = clean_company_name(cbx_row[CBX_COMPANY_OLD]) if cbx_row[CBX_COMPANY_OLD] else ''
        
        # Quick rejection filters (very fast string operations)
        # Filter 1: Smart first character match
        # - For normal names: check first character, but also check after removing common prefixes
        # - For acronyms: check if HC acronym appears anywhere in CBX name
        
        passes_first_char = False
        
        # Helper function to remove prefixes
        def remove_prefixes(name):
            for prefix in common_prefixes:
                if name.startswith(prefix):
                    return name[len(prefix):].strip()
            return name
        
        # Get all name variants (original and without prefixes)
        cbx_en_variants = [cbx_en, remove_prefixes(cbx_en)]
        cbx_fr_variants = [cbx_fr, remove_prefixes(cbx_fr)]
        cbx_old_variants = [cbx_old, remove_prefixes(cbx_old)] if cbx_old else []
        
        # Check direct first character match against all variants
        for variant in cbx_en_variants + cbx_fr_variants + cbx_old_variants:
            if variant and len(variant) > 0 and name_start[0] == variant[0]:
                passes_first_char = True
                break
        
        # For potential acronyms: check if acronym letters appear in order
        if not passes_first_char and is_likely_acronym:
            acronym = clean_name.replace(' ', '').lower()
            # Check if each letter of acronym appears in order in CBX name
            for cbx_name in [cbx_en, cbx_fr, cbx_old]:
                if not cbx_name:
                    continue
                # Try exact acronym match first
                if acronym in cbx_name.replace(' ', ''):
                    passes_first_char = True
                    break
                # Try fuzzy acronym: check if letters appear in sequence
                words = cbx_name.split()
                if len(words) >= len(acronym):
                    # Check if first letters of words form the acronym
                    first_letters = ''.join([w[0] for w in words if w])
                    if acronym in first_letters or first_letters.startswith(acronym):
                        passes_first_char = True
                        break
        
        if not passes_first_char:
            continue
        
        # Filter 2: Length similarity (companies with very different name lengths rarely match)
        # Allow +/- 20 characters difference (increased from 15 to handle more prefix variations)
        lengths = [len(cbx_en), len(cbx_fr)]
        if cbx_old:
            lengths.append(len(cbx_old))
        
        min_len = min(lengths)
        max_len = max(lengths)
        
        # Be more lenient for short names (they often have longer variations)
        tolerance = 30 if name_length <= 10 else 20
        
        if (name_length < min_len - tolerance or name_length > max_len + tolerance):
            continue
        
        # Filter 3: At least one common word (very discriminative)
        en_words = set(cbx_en.split())
        fr_words = set(cbx_fr.split())
        old_words = set(cbx_old.split()) if cbx_old else set()
        
        has_common_word = bool(hc_words & en_words or hc_words & fr_words or hc_words & old_words)
        
        # Must have at least one word in common with any CBX name variant
        if not has_common_word:
            # Exception 1: if names are very short (< 8 chars), be more lenient
            # Exception 2: if this might be an acronym match
            if name_length >= 8 and not is_likely_acronym:
                continue
        
        # Passed all filters - keep this candidate
        filtered.append(cbx_row)
    
    return filtered

def normalize_address(addr):
    import unicodedata
    addr = addr.lower()
    addr = unicodedata.normalize('NFKD', addr)
    addr = ''.join([c for c in addr if not unicodedata.combining(c)])  # Remove accents
    addr = re.sub(r'[.,\-]', ' ', addr)  # Remove punctuation and hyphens
    # Convert French street types/directions to English and handle abbreviations
    translations = {
        'chemin': 'road',
        'rue': 'street',
        'route': 'road',
        'boulevard': 'boulevard',
        'blvd': 'boulevard',
        'avenue': 'avenue',
        'place': 'place',
        'terrasse': 'terrace',
        'ouest': 'west',
        'o': 'west',
        'est': 'east',
        'e': 'east',
        'nord': 'north',
        'n': 'north',
        'sud': 'south',
        's': 'south',
        'appartement': 'apartment',
        'app': 'apartment',
        'batiment': 'building',
        'bâtiment': 'building',
        'immeuble': 'building',
        'etage': 'floor',
        'étage': 'floor',
        'suite': 'suite',
        'bureau': 'office',
    }
    for fr, en in translations.items():
        addr = re.sub(rf'\b{fr}\b', en, addr)
    addr = re.sub(r'\s+', ' ', addr)
    addr = addr.strip()
    # Remove duplicate words and sort for order-insensitive comparison
    words = sorted(set(addr.split()))
    return ' '.join(words)


def parse_assessment_level(level):
    if(level is None or (isinstance(level, int) and level > 0 and level < 4)):
        return level
    
    if(level.lower() in assessment_levels):
        return assessment_levels[level.lower()]
    
    return 0

if __name__ == '__main__':
    data_path = './data/'
    cbx_file = data_path + args.cbx_list
    hc_file = data_path + args.hc_list
    output_file = data_path + args.output

    # output parameters used
    print(f'Starting at {datetime.now()}')
    print(f'Reading CBX list: {args.cbx_list} [{args.cbx_encoding}]')
    print(f'Reading HC list: {args.hc_list}')
    print(f'Outputting results in: {args.output}')
    print(f'contractor match ratio: {args.ratio_company}')
    print(f'address match ratio: {args.ratio_address}')
    print(f'list of generic domains:\n{BASE_GENERIC_DOMAIN}')
    print(f'additional generic domain: {args.additional_generic_domain}')
    # read data
    cbx_data = []
    hc_data = []
    hc_row = []
    print('Reading Cognibox data file...')
    with open(cbx_file, 'r', encoding=args.cbx_encoding) as cbx:
        for row in csv.reader(cbx):
            cbx_data.append(row)
    # check cbx db ata consistency
    if cbx_data and len(cbx_data[0]) != len(cbx_headers):
        print(f'WARNING: got {len(cbx_data[0])} columns when expecting {len(cbx_headers)}')
        if not args.ignore_warnings:
            exit(-1)
    if not args.no_headers:
        headers = cbx_data.pop(0)
        headers = [x.lower().strip() for x in headers]
        check_headers(headers, cbx_headers, args.ignore_warnings)
    # for index, row in enumerate(cbx_data):
    #     access_modes = row[CBX_ACCESS_MODES].split(';')
    #     # only keep contractors on Non-member without any access mode (ignore training and hiring clients)
    #     if 'Contractor' not in access_modes and access_modes:
    #         cbx_data.pop(index)
    print(f'Completed reading {len(cbx_data)} contractors.')

    print('Reading hiring client data file...')
    hc_wb = openpyxl.load_workbook(hc_file, read_only=True, data_only=True)
    if args.hc_list_sheet_name:
        hc_sheet = hc_wb.get_sheet_by_name(args.hc_list_sheet_name)
    else:
        hc_sheet = hc_wb.active
    max_row = hc_sheet.max_row
    max_column = hc_sheet.max_column
    row_offset = 0 if not args.hc_list_offset else int(args.hc_list_offset.split(',')[0])-1
    column_offset = 0 if not args.hc_list_offset else int(args.hc_list_offset.split(',')[1])-1

    if max_column > 250 or max_row > 10000:
        print(f'WARNING: File is large: {max_row} rows and {max_column}. must be less than 10000 and 250')
        if not args.ignore_warnings:
            exit(-1)
    for row in hc_sheet.rows:
        # start data retrieval at offset
        while row_offset:
            next(hc_sheet.rows)
            row_offset -= 1
        row = row[column_offset:]
        # retrieve
        if not row[0].value:
            continue
        hc_data.append([cell.value if cell.value is not None else '' for cell in row])
    total = len(hc_data) - 1
    metadata_indexes = []
    headers = []
    rd_headers_mapping = []
    hs_headers_mapping = []
    existing_contractors_headers_mapping = []
    # check hc data consistency
    if hc_data and len(hc_data[0]) < len(hiring_client_headers):
        print(f'WARNING: got {len(hc_data[0])} columns when at least {len(hiring_client_headers)} is expected')
        if not args.ignore_warnings:
            exit(-1)
    if not args.no_headers:
        headers = hc_data.pop(0)
        headers = [x.lower().strip() for x in headers]
        check_headers(headers, hiring_client_headers, args.ignore_warnings)
    else:
        if hc_data and len(hc_data[0]) != len(hiring_client_headers):
            print(f'WARNING: got {len(hc_data[0])} columns when {len(hiring_client_headers)} is exactly expected')
            if not args.ignore_warnings:
                exit(-1)
    # checking currency integrity and strip characters from contact phone
    for row in hc_data:
            # Ignore extra columns: only process expected columns, leave extras untouched
            # Trim whitespace from all fields
            row = [str(cell).strip() if cell is not None else '' for cell in row]
            # Ensure company name and address are UTF-8 encoded and normalized
            if row[HC_COMPANY]:
                row[HC_COMPANY] = row[HC_COMPANY].encode('utf-8', errors='ignore').decode('utf-8')
            if row[HC_STREET]:
                row[HC_STREET] = row[HC_STREET].encode('utf-8', errors='ignore').decode('utf-8')
            # Existing normalization logic
            if row[HC_COUNTRY].lower().strip() == 'ca':
                if row[HC_CONTACT_CURRENCY].lower().strip() not in ('cad', ''):
                    print(f'WARNING: currency and country mismatch: {row[HC_CONTACT_CURRENCY]} and'
                          f' "{row[HC_COUNTRY]}". Expected CAD in row {row}')
                    if not args.ignore_warnings:
                        exit(-1)
            elif row[HC_COUNTRY].lower().strip() != '':
                if row[HC_CONTACT_CURRENCY].lower().strip() not in ('usd', ''):
                    print(f'WARNING: currency and country mismatch: {row[HC_CONTACT_CURRENCY]} and'
                          f' "{row[HC_COUNTRY]}". Expected USD in row {row}')
                    if not args.ignore_warnings:
                        exit(-1)
            row[HC_EMAIL] = str(row[HC_EMAIL]).strip()
            # correct and normalize phone number
            extension = ''
            if isinstance(row[HC_CONTACT_PHONE], str):
                for x in ('ext', 'x', 'poste', ',', 'p'):
                    f_index = row[HC_CONTACT_PHONE].lower().find(x)
                    if f_index >= 0:
                        extension = row[HC_CONTACT_PHONE][f_index + len(x):]
                        row[HC_CONTACT_PHONE] = row[HC_CONTACT_PHONE][0:f_index]
                        break
                row[HC_CONTACT_PHONE] = re.sub("[^0-9]", "", row[HC_CONTACT_PHONE])
            elif isinstance(row[HC_CONTACT_PHONE], int):
                row[HC_CONTACT_PHONE] = str(row[HC_CONTACT_PHONE])
            if row[HC_CONTACT_PHONE] and not row[HC_PHONE]:
                row[HC_PHONE] = row[HC_CONTACT_PHONE]
                row[HC_EXTENSION] = extension
            if isinstance(row[HC_EXTENSION], str):
                row[HC_EXTENSION] = re.sub("[^0-9]", "", row[HC_EXTENSION])
            # make language lower case; currency, state ISO2 and country ISO2 upper case
            row[HC_LANGUAGE] = row[HC_LANGUAGE].lower()
            row[HC_CONTACT_LANGUAGE] = row[HC_CONTACT_LANGUAGE].lower()
            row[HC_COUNTRY] = row[HC_COUNTRY].upper()
            row[HC_STATE] = row[HC_STATE].upper()
            row[HC_CONTACT_CURRENCY] = row[HC_CONTACT_CURRENCY].upper()
            # convert date-time to windows format
            row[HC_CONTACT_TIMEZONE] = convertFromIANATimezone(row[HC_CONTACT_TIMEZONE])
    print(f'Completed reading {len(hc_data)} contractors.')
    print(f'Starting data analysis...')

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'all'
    out_ws_onboarding = out_wb.create_sheet(title="onboarding")
    out_ws_association_fee = out_wb.create_sheet(title="association_fee")
    out_ws_re_onboarding = out_wb.create_sheet(title="re_onboarding")
    out_ws_subscription_upgrade = out_wb.create_sheet(title="subscription_upgrade")
    out_ws_ambiguous_onboarding = out_wb.create_sheet(title="ambiguous_onboarding")
    out_ws_restore_suspended = out_wb.create_sheet(title="restore_suspended")
    out_ws_activation_link = out_wb.create_sheet(title="activation_link")
    out_ws_already_qualified = out_wb.create_sheet(title="already_qualified")
    out_ws_add_questionnaire = out_wb.create_sheet(title="add_questionnaire")
    out_ws_missing_information = out_wb.create_sheet(title="missing_info")
    out_ws_follow_up_qualification = out_wb.create_sheet(title="follow_up_qualification")
    out_ws_onboarding_rd = out_wb.create_sheet(title="Data to import")
    out_ws_existing_contractors = out_wb.create_sheet(title="Existing Contractors")
    out_ws_onboarding_hs = out_wb.create_sheet(title="Data for HS")

    sheets = (out_ws, out_ws_onboarding, out_ws_association_fee, out_ws_re_onboarding, out_ws_subscription_upgrade,
              out_ws_ambiguous_onboarding, out_ws_restore_suspended, out_ws_activation_link, out_ws_already_qualified,
              out_ws_add_questionnaire, out_ws_missing_information, out_ws_follow_up_qualification,
              out_ws_onboarding_rd, out_ws_existing_contractors, out_ws_onboarding_hs)

    # append analysis headers and move metadata headers at the end
    if not args.no_headers:
        for idx, val in enumerate(headers):
            if val.lower().startswith('metadata'):
                metadata_indexes.append(idx)
        metadata_indexes.sort(reverse=True)
        headers.extend(analysis_headers)
        metadata_array = []
        for md_index in metadata_indexes:
            metadata_array.insert(0, headers.pop(md_index))
        headers.extend(metadata_array)
        hubspot_headers.extend(metadata_array)  # hubspot headers must includes metadata if present
        existing_contractors_headers.extend(metadata_array)  # existing contractors headers must includes metadata if present
        rd_headers.extend(metadata_array)
        column_rd = column_hs = column_existing_contractors = 0
        for index, value in enumerate(headers):
            # skip the last two sheets since they have special mapping handled below
            for sheet in sheets[:-3]:
                sheet.cell(1, index+1, value)
            rd_headers_for_value = [s for s in rd_headers if value in s]
            if rd_headers_for_value:
                column_rd += 1
                rd_headers_mapping.append(True)
                # Invert code and id columns
                if value == "pricing_group_id":
                    adjustement = 1
                    rd_pricing_group_id_col = column_rd
                elif value == "pricing_group_code":
                    adjustement = -1
                    rd_pricing_group_code_col = column_rd
                else:
                    adjustement = 0

                if value in rd_headers:
                    out_ws_onboarding_rd.cell(1, column_rd + adjustement, value)
                else:
                    out_ws_onboarding_rd.cell(1, column_rd, rd_headers_for_value[0])
            else:
                rd_headers_mapping.append(False)

            if value in hubspot_headers:
                column_hs += 1
                hs_headers_mapping.append(True)
                out_ws_onboarding_hs.cell(1, column_hs, value)
            else:
                hs_headers_mapping.append(False)

            existing_contractors_headers_for_value = [s for s in existing_contractors_headers if value in s]
            if existing_contractors_headers_for_value:
                column_existing_contractors += 1
                existing_contractors_headers_mapping.append(True)
                if value in existing_contractors_headers:
                    out_ws_existing_contractors.cell(1, column_existing_contractors, value)
                else:
                    out_ws_existing_contractors.cell(1, column_existing_contractors, existing_contractors_headers_for_value[0])
            else:
                existing_contractors_headers_mapping.append(False)
                
        out_wb.save(filename=output_file)
    # match
    for index, hc_row in enumerate(hc_data):
        print(f"Processing {index+1}/{len(hc_data)}: {hc_row[HC_COMPANY]}")
        
        # Initialize variables for this iteration
        best_match = None
        analysis_string = ''
        is_ambiguous = False
        selected_candidate = None
        best_ratio_company = 0
        best_ratio_address = 0
        
        # CRITICAL: Check HC_DO_NOT_MATCH flag - if True, check mandatory fields first, then try company matching
        if smart_boolean(hc_row[HC_DO_NOT_MATCH]):
            print(f"  -> HC_DO_NOT_MATCH flag set - checking mandatory fields first")
            
            # Check if all mandatory fields are present
            if core_mandatory_provided(hc_row):
                print(f"     All mandatory fields present - marking as onboarding")
                wave = 'onboarding'
                
                # Build output row with empty analysis columns (no CBX matching needed for complete records)
                output_row = hc_row[:HC_HEADER_LENGTH]
                for _ in analysis_headers:
                    output_row.append('')
                analysis_string = 'Complete HC record - no CBX matching required'
            else:
                print(f"     Missing mandatory fields - attempting company-only fuzzy matching")
                
                # Try fuzzy company matching even with missing fields
                hc_company = hc_row[HC_COMPANY]
                clean_hc_company = clean_company_name(hc_company)
                
                best_company_match = None
                best_company_score = 0
                company_analysis = f"Company-only matching attempt for: {hc_company}\n"
                
                # Simple company fuzzy matching loop
                for cbx_row in cbx_data:
                    cbx_company_en = clean_company_name(cbx_row[CBX_COMPANY_EN])
                    cbx_company_fr = clean_company_name(cbx_row[CBX_COMPANY_FR])
                    
                    # Test both English and French company names
                    for cbx_company in [cbx_company_en, cbx_company_fr]:
                        if cbx_company.strip():
                            # Use same fuzzy matching logic as main algorithm
                            from fuzzywuzzy import fuzz
                            ratio = fuzz.ratio(clean_hc_company, cbx_company)
                            token_sort_ratio = fuzz.token_sort_ratio(clean_hc_company, cbx_company)
                            token_set_ratio = fuzz.token_set_ratio(clean_hc_company, cbx_company)
                            partial_ratio = fuzz.partial_ratio(clean_hc_company, cbx_company)
                            
                            # Use the maximum of all ratios (same as main algorithm)
                            max_ratio = max(ratio, token_sort_ratio, token_set_ratio, partial_ratio)
                            
                            if max_ratio > best_company_score:
                                best_company_score = max_ratio
                                best_company_match = cbx_row
                                company_analysis += f"Best match: {cbx_row[CBX_COMPANY_EN]} (Score: {max_ratio})\n"
                
                # If we found a good company match (>= 70), proceed with action logic
                if best_company_score >= 70 and best_company_match:
                    print(f"     Found company match: {best_company_match[CBX_COMPANY_EN]} (Score: {best_company_score})")
                    
                    # Create match_data dictionary similar to main matching logic
                    match_data = {
                        'company': best_company_match[CBX_COMPANY_EN],
                        'email': best_company_match[CBX_EMAIL],
                        'first_name': best_company_match[CBX_FISTNAME],  # Note: typo in original CBX_FISTNAME
                        'last_name': best_company_match[CBX_LASTNAME],
                        'registration_status': best_company_match[CBX_REGISTRATION_STATUS],
                        'expiration_date': best_company_match[CBX_EXPIRATION_DATE],
                        'is_qualified': False,  # Default to False for company-only matches
                        'ratio_company': best_company_score,
                        'ratio_address': 0
                    }
                    
                    # Parse expiration date
                    expiration_date = None
                    if best_company_match[CBX_EXPIRATION_DATE]:
                        try:
                            expiration_date = datetime.strptime(best_company_match[CBX_EXPIRATION_DATE], '%Y-%m-%d')
                        except (ValueError, TypeError):
                            expiration_date = None
                    
                    # Determine action using normal business logic
                    create = True  # Missing contact info means new contractor
                    wave = action(hc_row, match_data, create, False, expiration_date, False, args.ignore_warnings)
                    
                    # Build analysis with company match details
                    analysis_string = company_analysis + f"\nMatched CBX record: {best_company_match[CBX_COMPANY_EN]}\n"
                    analysis_string += f"Registration Status: {best_company_match[CBX_REGISTRATION_STATUS]}\n"
                    analysis_string += f"Contact: {best_company_match[CBX_FISTNAME]} {best_company_match[CBX_LASTNAME]} ({best_company_match[CBX_EMAIL]})\n"
                    analysis_string += f"Company Match Score: {best_company_score}\n"
                    analysis_string += f"NOTE: Missing contact info - proceeding with company match only"
                    
                    # Build output row with populated analysis
                    output_row = hc_row[:HC_HEADER_LENGTH]
                    for header in analysis_headers:
                        if header == 'analysis':
                            output_row.append(analysis_string)
                        elif header == 'cbx_company_en':
                            output_row.append(best_company_match[CBX_COMPANY_EN])
                        elif header == 'cbx_company_fr':
                            output_row.append(best_company_match[CBX_COMPANY_FR])
                        elif header == 'cbx_contact_first':
                            output_row.append(best_company_match[CBX_FISTNAME])
                        elif header == 'cbx_contact_last':
                            output_row.append(best_company_match[CBX_LASTNAME])
                        elif header == 'cbx_contact_email':
                            output_row.append(best_company_match[CBX_EMAIL])
                        elif header == 'name_score':
                            output_row.append(best_company_score)
                        elif header == 'address_score':
                            output_row.append(0)  # No address matching for company-only
                        else:
                            output_row.append('')
                else:
                    print(f"     No good company match found (best score: {best_company_score}) - marking as missing_info")
                    wave = 'missing_info'
                    
                    # Build output row with empty analysis columns
                    output_row = hc_row[:HC_HEADER_LENGTH]
                    for _ in analysis_headers:
                        output_row.append('')
                    analysis_string = company_analysis + "No suitable company match found"
            
            # Set action and index with safe column access
            if 'action' in analysis_headers:
                output_row[HC_HEADER_LENGTH + analysis_headers.index('action')] = wave
            if 'index' in analysis_headers:
                output_row[HC_HEADER_LENGTH + analysis_headers.index('index')] = index + 1
        
            # Add metadata columns if they exist
            if metadata_indexes:
                metadata_array = []
                for md_index in metadata_indexes:
                    metadata_array.insert(0, hc_row[md_index])
                output_row += metadata_array

            # Save the processed row and continue to next contractor
            hc_data[index] = output_row
            out_ws.append(output_row)
            continue
        
        # PRIORITY CHECK: Handle force_cbx_id first (manual override)
        force_cbx_id = None
        if hc_row[HC_FORCE_CBX_ID] and str(hc_row[HC_FORCE_CBX_ID]).strip() not in ('', 'nan', 'None'):
            try:
                force_cbx_id = int(float(str(hc_row[HC_FORCE_CBX_ID]).strip()))
                print(f"  -> Force CBX ID specified: {force_cbx_id}")
            except (ValueError, TypeError):
                print(f"  -> Invalid force_cbx_id value: {hc_row[HC_FORCE_CBX_ID]}")
                force_cbx_id = None
        
        # If force_cbx_id is specified, find that specific CBX contractor
        if force_cbx_id:
            forced_match = None
            for cbx_row in cbx_data:
                # Convert CBX_ID to int for comparison (CSV data is loaded as strings)
                try:
                    cbx_id = int(cbx_row[CBX_ID])
                    if cbx_id == force_cbx_id:
                        forced_match = cbx_row
                        break
                except (ValueError, TypeError):
                    continue  # Skip invalid CBX_ID entries
            
            if forced_match:
                print(f"  -> Found forced CBX match: {forced_match[CBX_COMPANY_EN]} (ID: {force_cbx_id})")
                # Use forced match and skip fuzzy matching
                best_match = forced_match
                analysis_string = f"FORCED MATCH: CBX ID {force_cbx_id} - {forced_match[CBX_COMPANY_EN]}\n"
                is_ambiguous = smart_boolean(hc_row[HC_AMBIGUOUS])
                # Set selected_candidate and ratio values for forced matches
                selected_candidate = {
                    'match': forced_match,
                    'name_score': 100,  # Perfect match since it was forced
                    'address_score': 100,
                    'contact_match': False
                }
                best_ratio_company = 100  # Perfect match since it was forced
                best_ratio_address = 100
            else:
                print(f"  -> WARNING: Forced CBX ID {force_cbx_id} not found in CBX database")
                best_match = None
                analysis_string = f"FORCED MATCH FAILED: CBX ID {force_cbx_id} not found\n"
                is_ambiguous = False
                selected_candidate = None
                best_ratio_company = 0
                best_ratio_address = 0
        else:
            # No forced ID - proceed with normal fuzzy matching
            hc_company = hc_row[HC_COMPANY]
            clean_hc_company = clean_company_name(hc_company)
            hc_zip = normalize_postal_code(hc_row[HC_ZIP])
            
            hc_address = str(hc_row[HC_STREET]).lower().replace('.', '').replace(',', '').replace('north', 'n').replace('south', 's').replace('east', 'e').replace('west', 'w').replace('  ', ' ').strip()
            hc_address_norm = normalize_address(hc_address)
            suite_keywords = ['suite', 'bureau', 'office']
            hc_suite = [w for w in hc_address_norm.split() if w in suite_keywords]
            candidates = []
            analysis_string = ''
            best_name_score = -1
            
            # Fuzzy matching loop (only when no forced match)
            for cbx_row in cbx_data:
                cbx_company_en = clean_company_name(cbx_row[CBX_COMPANY_EN])
                cbx_company_fr = clean_company_name(cbx_row[CBX_COMPANY_FR])
                ratio_company_fr = max(
                    fuzz.token_sort_ratio(cbx_company_fr, clean_hc_company),
                    fuzz.partial_ratio(cbx_company_fr, clean_hc_company),
                    fuzz.token_set_ratio(cbx_company_fr, clean_hc_company)
                )
                ratio_company_en = max(
                    fuzz.token_sort_ratio(cbx_company_en, clean_hc_company),
                    fuzz.partial_ratio(cbx_company_en, clean_hc_company),
                    fuzz.token_set_ratio(cbx_company_en, clean_hc_company)
                )
                ratio_company = max(ratio_company_fr, ratio_company_en)
                cbx_zip = normalize_postal_code(cbx_row[CBX_ZIP])
                cbx_address = normalize_address(cbx_row[CBX_ADDRESS])
                cbx_suite = [w for w in cbx_address.split() if w in suite_keywords]
                ratio_address = max(
                    fuzz.token_sort_ratio(cbx_address, hc_address_norm),
                    fuzz.partial_ratio(cbx_address, hc_address_norm),
                    fuzz.token_set_ratio(cbx_address, hc_address_norm)
                )
                zip_match_strict = hc_zip and cbx_zip and hc_zip == cbx_zip
                suite_match_strict = hc_suite and cbx_suite and bool(set(hc_suite) & set(cbx_suite))
                address_threshold = min(float(args.ratio_address), 80)
                address_match = ratio_address >= address_threshold
                name_substring = clean_hc_company.lower() in cbx_company_en.lower() or clean_hc_company.lower() in cbx_company_fr.lower() or cbx_company_en.lower() in clean_hc_company.lower() or cbx_company_fr.lower() in clean_hc_company.lower()
                # Also check old names
                cbx_old_names = clean_company_name(cbx_row[CBX_COMPANY_OLD]) if cbx_row[CBX_COMPANY_OLD] else ''
                if cbx_old_names:
                    ratio_company_old = max(
                        fuzz.token_sort_ratio(cbx_old_names, clean_hc_company),
                        fuzz.partial_ratio(cbx_old_names, clean_hc_company),
                        fuzz.token_set_ratio(cbx_old_names, clean_hc_company)
                    )
                    ratio_company = max(ratio_company, ratio_company_old)
                    name_substring = name_substring or clean_hc_company.lower() in cbx_old_names.lower() or cbx_old_names.lower() in clean_hc_company.lower()
                if (ratio_company >= float(args.ratio_company) or name_substring):
                    hiring_clients_list = cbx_row[CBX_HIRING_CLIENT_NAMES].split(args.list_separator) if cbx_row[CBX_HIRING_CLIENT_NAMES] else []
                    hc_count = len(hiring_clients_list)
                    contact_match = hc_row[HC_EMAIL] == cbx_row[CBX_EMAIL] and hc_row[HC_FIRSTNAME].lower() == cbx_row[CBX_FISTNAME].lower() and hc_row[HC_LASTNAME].lower() == cbx_row[CBX_LASTNAME].lower()
                    
                    # Calculate business value factors
                    module_count = len(cbx_row[CBX_MODULES].split(';')) if cbx_row[CBX_MODULES] and cbx_row[CBX_MODULES].strip() else 0
                    
                    # Calculate location proximity bonus (including country matching/penalties)
                    location_bonus = calculate_location_bonus(
                        hc_row[HC_STREET], hc_row[HC_CITY], hc_row[HC_STATE],
                        cbx_row[CBX_ADDRESS], cbx_row[CBX_CITY], cbx_row[CBX_STATE],
                        hc_row[HC_COUNTRY], cbx_row[CBX_COUNTRY]
                    )
                    
                    # Calculate business quality score (higher is better)
                    # FIXED: Reduced HC count weight from 2 to 0.5 to prevent high-volume contractors from dominating selection
                    # Prioritize verification and location accuracy over business volume
                    business_score = (
                        module_count * 3 +                      # Modules: 3 points each (reduced from 5, qualification breadth)
                        min(hc_count * 0.5, 5) +               # Hiring clients: 0.5 points each, capped at 5 total (reduced dominance)
                        (30 if contact_match else 0) +         # Contact match: 30 point bonus (increased from 20, verification priority)
                        (25 if zip_match_strict else 0) +      # Postal code match: 25 point bonus (increased from 10, location accuracy)
                        (15 if suite_match_strict else 0) +    # Suite match: 15 point bonus (increased from 5, address precision)
                        location_bonus * 2                     # Location proximity: doubled weight for geographical relevance
                    )
                    
                    candidates.append({
                        'cbx_row': cbx_row,
                        'name_score': ratio_company,
                        'address_score': ratio_address,
                        'business_score': business_score,
                        'module_count': module_count,
                        'hc_count': hc_count,
                        'contact_match': contact_match,
                        'location_bonus': location_bonus,
                        'zip': cbx_zip,
                        'suite': cbx_suite,
                        'zip_match_strict': zip_match_strict,
                        'suite_match_strict': suite_match_strict
                    })
                    if ratio_company >= 50 or ratio_address >= 50 or name_substring:
                        # Only include very close matches in analysis (90%+ company match)
                        # Also ensure CBX entry has meaningful data (not mostly empty)
                        cbx_company = cbx_row[CBX_COMPANY_EN] or cbx_row[CBX_COMPANY_FR] or ""
                        cbx_has_meaningful_data = (
                            cbx_company.strip() and len(cbx_company.strip()) > 2 and  # Company name exists and is not too short
                            cbx_company.lower().strip() not in ('main department', 'ontario', 'montreal', 'ver', '')  # Filter out generic entries
                        ) or (
                            cbx_row[CBX_ADDRESS] or cbx_row[CBX_EMAIL] or cbx_row[CBX_FISTNAME]  # Has address or contact info
                        )
                        if ratio_company >= 70 and cbx_has_meaningful_data:
                            analysis_string += f"{cbx_row[CBX_ID]}, {cbx_row[CBX_COMPANY_EN]}, {cbx_row[CBX_ADDRESS]}, {cbx_row[CBX_CITY]}, {cbx_row[CBX_STATE]}, {cbx_row[CBX_ZIP]}, {cbx_row[CBX_COUNTRY]}, {cbx_row[CBX_EMAIL]}, {cbx_row[CBX_FISTNAME]} {cbx_row[CBX_LASTNAME]} --> CR{ratio_company}, AR{ratio_address}, CM{contact_match}, HCC{hc_count}, M[{cbx_row[CBX_MODULES]}]\n"
                    if ratio_company > best_name_score:
                        best_name_score = ratio_company
            
            # Enhanced priority-based selection with business logic
            selected_candidate = None
            
            if candidates:
                # Only consider candidates that meet analysis threshold (company >= 70 and meaningful data)
                eligible_candidates = [c for c in candidates if c['name_score'] >= 70]
                
                if eligible_candidates:
                    # Priority 1: Perfect company name matches (>=95) - prioritize relationships first
                    perfect_matches = [c for c in eligible_candidates if c['name_score'] >= 95]
                    if perfect_matches:
                        # FIXED: Prioritize hiring client relationships above all else, then verification and location
                        # Sort by: hc_count (relationship depth), contact_match, zip_match, address_score, business_score
                        perfect_matches = sorted(perfect_matches, key=lambda c: (
                            int(c['hc_count'] > 0),       # Relationship exists (1 > 0)
                            int(c['hc_count']),           # Relationship depth (more clients = stronger relationship)
                            int(c['contact_match']) if isinstance(c['contact_match'], bool) else (1 if str(c['contact_match']).lower() == 'true' else 0),  # Contact verification 
                            int(c['zip_match_strict']) if isinstance(c['zip_match_strict'], bool) else (1 if str(c['zip_match_strict']).lower() == 'true' else 0),  # Postal code accuracy
                            float(c['address_score']) if c['address_score'] is not None else 0.0,  # Address similarity
                            float(c['business_score']) if c['business_score'] is not None else 0.0,  # Business factors
                            float(c['name_score']) if c['name_score'] is not None else 0.0  # Name score
                        ), reverse=True)
                        selected_candidate = perfect_matches[0]
                    else:
                        # Priority 2: High company name matches (>=90) - prioritize relationships
                        high_company_matches = [c for c in eligible_candidates if c['name_score'] >= 90]
                        if high_company_matches:
                            # FIXED: Prioritize relationships even for high matches
                            high_company_matches = sorted(high_company_matches, key=lambda c: (
                                int(c['hc_count'] > 0),   # Relationship exists
                                int(c['hc_count']),       # Relationship depth
                                int(c['contact_match']) if isinstance(c['contact_match'], bool) else (1 if str(c['contact_match']).lower() == 'true' else 0),  # Contact verification
                                float(c['business_score']) if c['business_score'] is not None else 0.0,  # Business factors
                                float(c['name_score']) if c['name_score'] is not None else 0.0,  # Company match quality
                                float(c['address_score']) if c['address_score'] is not None else 0.0  # Address match quality
                            ), reverse=True)
                            selected_candidate = high_company_matches[0]
                        else:
                            # Priority 3: Good matches (>=70) with postal code - business factors + location accuracy
                            good_company_postal = [c for c in eligible_candidates if c['name_score'] >= 70 and hc_zip and c['zip'] == hc_zip]
                            if good_company_postal:
                                # Location + business factors for good matches
                                good_company_postal = sorted(good_company_postal, key=lambda c: (c['business_score'], c['name_score'], c['address_score']), reverse=True)
                                selected_candidate = good_company_postal[0]
                            else:
                                # Priority 4: Best overall combination - relationship-first approach
                                # FIXED: Always prioritize relationship depth, then combined scoring
                                for candidate in eligible_candidates:
                                    # Enhanced scoring that heavily weighs relationships
                                    relationship_bonus = candidate['hc_count'] * 20  # 20 points per hiring client relationship
                                    candidate['combined_score'] = (
                                        relationship_bonus +                    # Relationship depth (20 points per client)
                                        candidate['name_score'] * 0.3 +        # Company match (30%)
                                        candidate['business_score'] * 0.3 +     # Business factors (30%) 
                                        candidate['address_score'] * 0.2        # Address match (20%)
                                    )
                                combo_candidates = sorted(eligible_candidates, key=lambda c: (
                                    int(c['hc_count'] > 0),   # Relationship exists (boolean priority)
                                    float(c['combined_score']) if c['combined_score'] is not None else 0.0  # Then combined score
                                ), reverse=True)
                                selected_candidate = combo_candidates[0]
        
        # Extract best match data (common for both forced and fuzzy matching)
        if force_cbx_id:
            # For forced matches, we already have best_match set above
            best_ratio_company = 100  # Forced match = perfect score
            best_ratio_address = 100  # Forced match = perfect score
        elif selected_candidate:
            best_match = selected_candidate['cbx_row']
            best_ratio_company = selected_candidate['name_score']
            best_ratio_address = selected_candidate['address_score']
        else:
            best_match = None
            best_ratio_company = 0
            best_ratio_address = 0
        
        # Determine match quality (different logic for forced vs fuzzy matches)
        if best_match:
            if force_cbx_id:
                # For forced matches, is_ambiguous was already set during forced matching
                contact_match = hc_row[HC_EMAIL] == best_match[CBX_EMAIL] and hc_row[HC_FIRSTNAME].lower() == best_match[CBX_FISTNAME].lower() and hc_row[HC_LASTNAME].lower() == best_match[CBX_LASTNAME].lower()
            else:
                # For fuzzy matches, use the existing logic
                contact_match = hc_row[HC_EMAIL] == best_match[CBX_EMAIL] and hc_row[HC_FIRSTNAME].lower() == best_match[CBX_FISTNAME].lower() and hc_row[HC_LASTNAME].lower() == best_match[CBX_LASTNAME].lower()
                
                # Determine if this is a good match, ambiguous, or should be onboarding
                if best_ratio_company >= 70 and best_ratio_address < 70:  # Good company name but poor address
                    is_ambiguous = True
                elif best_ratio_company < 50:  # Very poor company name match
                    # This should be treated as no match/onboarding
                    best_match = None
        
        # Build analysis string if we have a match
        if best_match:
            # Only create single-entry analysis if no multi-entry analysis was built at all
            if not analysis_string.strip():
                hiring_clients_list = best_match[CBX_HIRING_CLIENT_NAMES].split(args.list_separator) if best_match[CBX_HIRING_CLIENT_NAMES] else []
                hc_count = len(hiring_clients_list)
                contact_match = hc_row[HC_EMAIL] == best_match[CBX_EMAIL] and hc_row[HC_FIRSTNAME].lower() == best_match[CBX_FISTNAME].lower() and hc_row[HC_LASTNAME].lower() == best_match[CBX_LASTNAME].lower()
                analysis_string = f"{best_match[CBX_ID]}, {best_match[CBX_COMPANY_EN]}, {best_match[CBX_ADDRESS]}, {best_match[CBX_CITY]}, {best_match[CBX_STATE]}, {best_match[CBX_ZIP]}, {best_match[CBX_COUNTRY]}, {best_match[CBX_EMAIL]}, {best_match[CBX_FISTNAME]} {best_match[CBX_LASTNAME]} --> CR{best_ratio_company}, AR{best_ratio_address}, CM{contact_match}, HCC{hc_count}, M[{best_match[CBX_MODULES]}]\n"
            
            # For good matches, add a header to show which entry was selected as the best match
            if not is_ambiguous and analysis_string.strip():
                # Prepend the selected match info to the analysis for clarity
                hiring_clients_list = best_match[CBX_HIRING_CLIENT_NAMES].split(args.list_separator) if best_match[CBX_HIRING_CLIENT_NAMES] else []
                hc_count = len(hiring_clients_list)
                # Calculate contact match for the selected candidate
                selected_contact_match = (hc_row[HC_EMAIL] == best_match[CBX_EMAIL] and 
                                        hc_row[HC_FIRSTNAME].lower() == best_match[CBX_FISTNAME].lower() and 
                                        hc_row[HC_LASTNAME].lower() == best_match[CBX_LASTNAME].lower()) if selected_candidate else False
                selected_match_info = f">>> SELECTED BEST MATCH: {best_match[CBX_ID]}, {best_match[CBX_COMPANY_EN]}, {best_match[CBX_ADDRESS]}, {best_match[CBX_CITY]}, {best_match[CBX_STATE]}, {best_match[CBX_ZIP]}, {best_match[CBX_COUNTRY]}, {best_match[CBX_EMAIL]}, {best_match[CBX_FISTNAME]} {best_match[CBX_LASTNAME]} --> CR{best_ratio_company}, AR{best_ratio_address}, CM{selected_contact_match}, HCC{hc_count}, M[{best_match[CBX_MODULES]}]\n\n>>> ALL CANDIDATES CONSIDERED:\n"
                analysis_string = selected_match_info + analysis_string
        
        # Process the contractor based on match quality
        if best_match and not is_ambiguous:
                # Good match - populate CBX columns and use action() for proper categorization
                # Calculate contact match
                contact_match = (hc_row[HC_EMAIL] == best_match[CBX_EMAIL] and 
                               hc_row[HC_FIRSTNAME].lower() == best_match[CBX_FISTNAME].lower() and 
                               hc_row[HC_LASTNAME].lower() == best_match[CBX_LASTNAME].lower())
                match_data = add_analysis_data(hc_row, best_match, analysis_string, best_ratio_company, best_ratio_address, contact_match)
                
                # Calculate is_qualified and expiration_date for action
                hiring_clients_list = best_match[CBX_HIRING_CLIENT_NAMES].split(args.list_separator) if best_match[CBX_HIRING_CLIENT_NAMES] else []
                hiring_clients_qstatus = best_match[CBX_HIRING_CLIENT_QSTATUS].split(args.list_separator) if best_match[CBX_HIRING_CLIENT_QSTATUS] else []
                is_qualified = False
                for idx, val in enumerate(hiring_clients_list):
                    if val == hc_row[HC_HIRING_CLIENT_NAME] and idx < len(hiring_clients_qstatus):
                        status = hiring_clients_qstatus[idx].lower().strip()
                        # Check for APPROVED/QUALIFIED statuses (as per user feedback)
                        if status in ('validated', 'validate', 'valid', 'approved', 'approve', 'qualified', 'active'):
                            is_qualified = True
                            break
                        # Handle NOT APPROVED/PENDING statuses  
                        elif status in ('not approved', 'not_approved', 'pending', 'pending approval', 'under review', 'in progress', 'expired', 'exprired'):
                            is_qualified = False
                            break
                
                # Safe date parsing with proper exception handling
                expiration_date = None
                if best_match[CBX_EXPIRATION_DATE]:
                    try:
                        expiration_date = datetime.strptime(best_match[CBX_EXPIRATION_DATE], "%d/%m/%y")
                    except ValueError:
                        try:
                            expiration_date = datetime.strptime(best_match[CBX_EXPIRATION_DATE], "%d/%m/%Y")
                        except ValueError:
                            print(f"WARNING: Could not parse expiration date: {best_match[CBX_EXPIRATION_DATE]}")
                            expiration_date = None
                
                # Calculate the action using the action() function
                create = False  # We found a match
                wave = action(hc_row, match_data, create, False, expiration_date, is_qualified, args.ignore_warnings)
                match_data['action'] = wave
                
                # Build output row with CBX data populated
                output_row = hc_row[:HC_HEADER_LENGTH]
                cbx_map = {
                    'cbx_contractor': 'company',
                    'cbx_street': 'address',
                    'cbx_city': 'city',
                    'cbx_state': 'state',
                    'cbx_zip': 'zip',
                    'cbx_country': 'country',
                    'cbx_expiration_date': 'expiration_date',
                    'cbx_email': 'email',
                    'cbx_first_name': 'first_name',
                    'cbx_last_name': 'last_name',
                    'cbx_account_type': 'account_type',
                    'cbx_subscription_fee': 'subscription_price',
                    'cbx_employee_price': 'employee_price',
                    'cbx_assessment_level': 'cbx_assessment_level',
                    'cbx_id': 'cbx_id',
                    'modules': 'modules',
                    'parents': 'parents',
                    'previous': 'previous',
                    'hiring_client_names': 'hiring_client_names',
                    'hiring_client_count': 'hiring_client_count',
                    'is_in_relationship': 'is_in_relationship',
                    'is_qualified': 'is_qualified',
                    'ratio_company': 'ratio_company',
                    'ratio_address': 'ratio_address',
                    'contact_match': 'contact_match',
                    'new_product': 'new_product',
                    'registration_status': 'registration_status',
                    'suspended': 'suspended',
                    'hc_contractor_summary': 'hc_contractor_summary',
                    'analysis': 'analysis'
                }
                for col_name in analysis_headers:
                    output_row.append(match_data.get(cbx_map.get(col_name, col_name), ''))
                
                # Set index for good matches - CRITICAL FIX
                if 'index' in analysis_headers:
                    output_row[HC_HEADER_LENGTH + analysis_headers.index('index')] = index + 1
                    
        elif best_match and is_ambiguous:
                # Ambiguous match - only populate CBX data if contractor has relationship with hiring client
                # Calculate contact match
                contact_match = (hc_row[HC_EMAIL] == best_match[CBX_EMAIL] and 
                               hc_row[HC_FIRSTNAME].lower() == best_match[CBX_FISTNAME].lower() and 
                               hc_row[HC_LASTNAME].lower() == best_match[CBX_LASTNAME].lower())
                match_data = add_analysis_data(hc_row, best_match, analysis_string, best_ratio_company, best_ratio_address, contact_match)
                
                # Check if this contractor has relationship with the hiring client
                has_relationship = match_data.get('is_in_relationship', False)
                
                # Set the ambiguous flag in hc_row for action() function
                temp_hc_row = hc_row[:]  # Create a copy
                if HC_AMBIGUOUS < len(temp_hc_row):
                    temp_hc_row[HC_AMBIGUOUS] = True
                
                # Only treat as existing contractor if they have a relationship
                create = not has_relationship  # If has relationship, treat as existing (create=False)
                wave = action(temp_hc_row, match_data, create, False, None, match_data.get('is_qualified', False), args.ignore_warnings)
                match_data['action'] = wave
                
                # Build output row - populate CBX data only if contractor has relationship
                output_row = hc_row[:HC_HEADER_LENGTH]
                
                if has_relationship:
                    # Full CBX data population for contractors with relationships
                    cbx_map = {
                        'cbx_id': 'cbx_id',
                        'hc_contractor_summary': 'hc_contractor_summary',
                        'analysis': 'analysis',
                        'cbx_contractor': 'cbx_contractor',
                        'cbx_street': 'cbx_street',
                        'cbx_city': 'cbx_city',
                        'cbx_state': 'cbx_state',
                        'cbx_zip': 'cbx_zip',
                        'cbx_country': 'cbx_country',
                        'cbx_expiration_date': 'cbx_expiration_date',
                        'registration_status': 'registration_status',
                        'suspended': 'suspended',
                        'cbx_email': 'cbx_email',
                        'cbx_first_name': 'cbx_first_name',
                        'cbx_last_name': 'cbx_last_name',
                        'modules': 'modules',
                        'cbx_account_type': 'cbx_account_type',
                        'cbx_subscription_fee': 'cbx_subscription_fee',
                        'cbx_employee_price': 'cbx_employee_price',
                        'parents': 'parents',
                        'previous': 'previous',
                        'hiring_client_names': 'hiring_client_names',
                        'hiring_client_count': 'hiring_client_count',
                        'is_in_relationship': 'is_in_relationship',
                        'is_qualified': 'is_qualified',
                        'ratio_company': 'ratio_company',
                        'ratio_address': 'ratio_address',
                        'contact_match': 'contact_match',
                        'cbx_assessment_level': 'cbx_assessment_level',
                        'new_product': 'new_product',
                        'generic_domain': 'generic_domain',
                        'match_count': 'match_count',
                        'match_count_with_hc': 'match_count_with_hc',
                        'is_subscription_upgrade': 'is_subscription_upgrade',
                        'upgrade_price': 'upgrade_price',
                        'prorated_upgrade_price': 'prorated_upgrade_price',
                        'create_in_cbx': 'create_in_cbx',
                        'action': 'action'
                    }
                    for col_name in analysis_headers:
                        output_row.append(match_data.get(cbx_map.get(col_name, col_name), ''))
                    # Set index for ambiguous matches
                    if 'index' in analysis_headers:
                        output_row[HC_HEADER_LENGTH + analysis_headers.index('index')] = index + 1
        else:
            # No match found - before marking as missing_info, try company-only matching
            print(f"  -> No fuzzy match found - attempting company-only matching for missing contact info")
            
            # Try company-only fuzzy matching similar to HC_DO_NOT_MATCH logic
            hc_company = hc_row[HC_COMPANY]
            clean_hc_company = clean_company_name(hc_company)
            
            best_company_match = None
            best_company_score = 0
            company_analysis = f"Company-only matching attempt for: {hc_company}\n"
            
            # Simple company fuzzy matching loop
            for cbx_row in cbx_data:
                cbx_company_en = clean_company_name(cbx_row[CBX_COMPANY_EN])
                cbx_company_fr = clean_company_name(cbx_row[CBX_COMPANY_FR])
                
                # Test both English and French company names
                for cbx_company in [cbx_company_en, cbx_company_fr]:
                    if cbx_company.strip():
                        # Use same fuzzy matching logic as main algorithm
                        from fuzzywuzzy import fuzz
                        ratio = fuzz.ratio(clean_hc_company, cbx_company)
                        token_sort_ratio = fuzz.token_sort_ratio(clean_hc_company, cbx_company)
                        token_set_ratio = fuzz.token_set_ratio(clean_hc_company, cbx_company)
                        partial_ratio = fuzz.partial_ratio(clean_hc_company, cbx_company)
                        
                        # Use the maximum of all ratios (same as main algorithm)
                        max_ratio = max(ratio, token_sort_ratio, token_set_ratio, partial_ratio)
                        
                        if max_ratio > best_company_score:
                            best_company_score = max_ratio
                            best_company_match = cbx_row
                            company_analysis += f"Best match: {cbx_row[CBX_COMPANY_EN]} (Score: {max_ratio})\n"
            
            # If we found a good company match (>= 70), proceed with action logic
            if best_company_score >= 70 and best_company_match:
                print(f"     Found company match: {best_company_match[CBX_COMPANY_EN]} (Score: {best_company_score})")
                
                # Create match_data dictionary similar to main matching logic
                match_data = {
                    'company': best_company_match[CBX_COMPANY_EN],
                    'email': best_company_match[CBX_EMAIL],
                    'first_name': best_company_match[CBX_FISTNAME],  # Note: typo in original CBX_FISTNAME
                    'last_name': best_company_match[CBX_LASTNAME],
                    'registration_status': best_company_match[CBX_REGISTRATION_STATUS],
                    'expiration_date': best_company_match[CBX_EXPIRATION_DATE],
                    'is_qualified': False,  # Default to False for company-only matches
                    'ratio_company': best_company_score,
                    'ratio_address': 0
                }
                
                # Parse expiration date
                expiration_date = None
                if best_company_match[CBX_EXPIRATION_DATE]:
                    try:
                        expiration_date = datetime.strptime(best_company_match[CBX_EXPIRATION_DATE], '%Y-%m-%d')
                    except (ValueError, TypeError):
                        expiration_date = None
                
                # Determine action using normal business logic
                create = True  # Missing contact info means new contractor
                wave = action(hc_row, match_data, create, False, expiration_date, False, args.ignore_warnings)
                
                # Build analysis with company match details
                analysis_string = company_analysis + f"\nMatched CBX record: {best_company_match[CBX_COMPANY_EN]}\n"
                analysis_string += f"Registration Status: {best_company_match[CBX_REGISTRATION_STATUS]}\n"
                analysis_string += f"Contact: {best_company_match[CBX_FISTNAME]} {best_company_match[CBX_LASTNAME]} ({best_company_match[CBX_EMAIL]})\n"
                analysis_string += f"Company Match Score: {best_company_score}\n"
                analysis_string += f"NOTE: Missing contact info - proceeding with company match only"
                
                # Build output row with populated analysis
                output_row = hc_row[:HC_HEADER_LENGTH]
                for header in analysis_headers:
                    if header == 'analysis':
                        output_row.append(analysis_string)
                    elif header == 'cbx_company_en':
                        output_row.append(best_company_match[CBX_COMPANY_EN])
                    elif header == 'cbx_company_fr':
                        output_row.append(best_company_match[CBX_COMPANY_FR])
                    elif header == 'cbx_contact_first':
                        output_row.append(best_company_match[CBX_FISTNAME])
                    elif header == 'cbx_contact_last':
                        output_row.append(best_company_match[CBX_LASTNAME])
                    elif header == 'cbx_contact_email':
                        output_row.append(best_company_match[CBX_EMAIL])
                    elif header == 'cbx_phone':
                        output_row.append('')  # No phone field in CBX data
                    elif header == 'name_score':
                        output_row.append(best_company_score)
                    elif header == 'address_score':
                        output_row.append(0)  # No address matching for company-only
                    else:
                        output_row.append('')
            else:
                print(f"     No good company match found (best score: {best_company_score}) - marking as missing_info")
                
                # No company match found - treat as create=True and use action() function
                create = True
                wave = action(hc_row, None, create, False, None, False, args.ignore_warnings)
                
                # Build output row with empty analysis columns
                output_row = hc_row[:HC_HEADER_LENGTH]
                for _ in analysis_headers:
                    output_row.append('')
            
            # Set action and index with safe column access
            if 'action' in analysis_headers:
                output_row[HC_HEADER_LENGTH + analysis_headers.index('action')] = wave
            if 'index' in analysis_headers:
                output_row[HC_HEADER_LENGTH + analysis_headers.index('index')] = index + 1
        
        # Add metadata columns if they exist
        if metadata_indexes:
            metadata_array = []
            for md_index in metadata_indexes:
                metadata_array.insert(0, hc_row[md_index])
            output_row += metadata_array
        
        # Save the processed row
        hc_data[index] = output_row
        out_ws.append(output_row)
            
    out_wb.save(filename=output_file)

    hc_onboarding = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'onboarding', hc_data)
    for index, row in enumerate(hc_onboarding):
        for i, value in enumerate(row):
            out_ws_onboarding.cell(index + 2, i + 1, value)

    hc_association_fee = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'association_fee', hc_data)
    for index, row in enumerate(hc_association_fee):
        for i, value in enumerate(row):
            out_ws_association_fee.cell(index + 2, i + 1, value)

    hc_re_onboarding = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 're_onboarding', hc_data)
    for index, row in enumerate(hc_re_onboarding):
        for i, value in enumerate(row):
            out_ws_re_onboarding.cell(index + 2, i + 1, value)

    hc_subscription_upgrade = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'subscription_upgrade',
                                     hc_data)
    for index, row in enumerate(hc_subscription_upgrade):
        for i, value in enumerate(row):
            out_ws_subscription_upgrade.cell(index + 2, i + 1, value)

    hc_ambiguous_onboarding = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'ambiguous_onboarding',
                                     hc_data)
    for index, row in enumerate(hc_ambiguous_onboarding):
        for i, value in enumerate(row):
            out_ws_ambiguous_onboarding.cell(index + 2, i + 1, value)

    hc_restore_suspended = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'restore_suspended',
                                  hc_data)
    for index, row in enumerate(hc_restore_suspended):
        for i, value in enumerate(row):
            out_ws_restore_suspended.cell(index + 2, i + 1, value)

    hc_activation_link = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'activation_link',
                                hc_data)
    for index, row in enumerate(hc_activation_link):
        for i, value in enumerate(row):
            out_ws_activation_link.cell(index + 2, i + 1, value)

    hc_already_qualified = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'already_qualified',
                                  hc_data)
    for index, row in enumerate(hc_already_qualified):
        for i, value in enumerate(row):
            out_ws_already_qualified.cell(index + 2, i + 1, value)

    hc_add_questionnaire = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'add_questionnaire',
                                  hc_data)
    for index, row in enumerate(hc_add_questionnaire):
        for i, value in enumerate(row):
            out_ws_add_questionnaire.cell(index + 2, i + 1, value)

    hc_missing_information = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'missing_info',
                                    hc_data)
    for index, row in enumerate(hc_missing_information):
        for i, value in enumerate(row):
            out_ws_missing_information.cell(index + 2, i + 1, value)

    hc_follow_up_qualification = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] ==
                                        'follow_up_qualification',
                                        hc_data)
    for index, row in enumerate(hc_follow_up_qualification):
        for i, value in enumerate(row):
            out_ws_follow_up_qualification.cell(index + 2, i + 1, value)

    existing_contractors_rd = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] != 'onboarding' and x[HC_HEADER_LENGTH+len(analysis_headers)-2] != 'missing_info' , hc_data)

    for index, row in enumerate(existing_contractors_rd):
        column = 0
        for i, value in enumerate(row):
            if existing_contractors_headers_mapping[i]:
                column += 1
                out_ws_existing_contractors.cell(index + 2, column, value)

    hc_onboarding_rd = filter(lambda x: x[HC_HEADER_LENGTH+len(analysis_headers)-2] == 'onboarding',
                              hc_data)
    for index, row in enumerate(hc_onboarding_rd):
        column = 0
        for i, value in enumerate(row):
            if rd_headers_mapping[i]:
                column += 1
                # Invert code and id columns
                if column == rd_pricing_group_id_col:
                    out_ws_onboarding_rd.cell(index + 2, column + 1, value)
                elif column == rd_pricing_group_code_col:
                    out_ws_onboarding_rd.cell(index + 2, column - 1, value)
                else:
                    out_ws_onboarding_rd.cell(index + 2, column, value)

    for index, row in enumerate(hc_data):
        column = 0
        for i, value in enumerate(row):
            if hs_headers_mapping[i]:
                column += 1
                out_ws_onboarding_hs.cell(index + 2, column, value)

    # formatting the excel...
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    dims = {}
    for sheet in sheets:
        tab = Table(displayName=sheet.title.replace(" ", "_"),
                    ref=f'A1:{get_column_letter(sheet.max_column)}{sheet.max_row + 1}')
        tab.tableStyleInfo = style
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value
        if sheet != out_ws_onboarding_rd:
            sheet.column_dimensions[get_column_letter(HC_HEADER_LENGTH+analysis_headers.index("hc_contractor_summary")+1)].width = 150
            sheet.column_dimensions[get_column_letter(HC_HEADER_LENGTH+analysis_headers.index("analysis")+1)].width = 150
            sheet.column_dimensions[get_column_letter(HC_HEADER_LENGTH+len(analysis_headers)-17)].width = 150
            sheet.column_dimensions[get_column_letter(HC_HEADER_LENGTH+len(analysis_headers)-18)].width = 150
            for i in range(2, len(hc_data)+1):
                sheet.cell(i, HC_HEADER_LENGTH+analysis_headers.index("analysis")+1).alignment = Alignment(wrapText=True)
                sheet.cell(i, HC_HEADER_LENGTH+analysis_headers.index("hc_contractor_summary")+1).alignment = Alignment(wrapText=True)
                sheet.cell(i, HC_HEADER_LENGTH+len(analysis_headers)-17).alignment = Alignment(wrapText=True)
                sheet.cell(i, HC_HEADER_LENGTH+len(analysis_headers)-18).alignment = Alignment(wrapText=True)
                
        sheet.add_table(tab)
    out_wb.save(filename=output_file)
    print(f'Completed data analysis...')
    print(f'Completed at {datetime.now()}')

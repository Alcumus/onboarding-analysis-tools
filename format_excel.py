#!/usr/bin/env python3
"""
Enhanced Excel Formatter for Parallel Processing Results
Only applies formatting (styles, tables, filters, column widths) - does NOT modify data
"""

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def apply_excel_formatting(input_file, output_file):
    """Apply Excel formatting (tables, styles, filters, column widths) to existing workbook"""
    
    print(f"Applying Excel formatting to {input_file} -> {output_file}")
    
    # Load the existing workbook (already has all sheets with data)
    wb = openpyxl.load_workbook(input_file)
    
    # Get all sheets
    sheets = wb.worksheets
    
    # Ensure all header cells are strings to avoid openpyxl warnings
    for sheet in sheets:
        if sheet.max_row >= 1:
            for cell in sheet[1]:  # First row (headers)
                if cell.value is not None:
                    cell.value = str(cell.value)
    
    # Get standard headers from the first non-empty sheet (usually 'all')
    standard_headers = None
    for sheet in sheets:
        if sheet.max_row >= 1 and sheet.max_column > 0:
            headers = [cell.value for cell in sheet[1]]
            if any(h is not None for h in headers):
                standard_headers = headers
                break
    
    # Add headers to empty sheets
    if standard_headers:
        for sheet in sheets:
            # Check if sheet is empty or has no headers
            if sheet.max_row < 1 or all(cell.value is None for cell in sheet[1]):
                print(f"  Adding headers to empty sheet: {sheet.title}")
                for col_idx, header in enumerate(standard_headers, start=1):
                    sheet.cell(row=1, column=col_idx, value=str(header) if header else "")
    
    # Apply formatting to all sheets
    style = TableStyleInfo(
        name="TableStyleMedium2", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    
    for sheet in sheets:
        # Skip completely empty sheets (shouldn't happen now that we add headers)
        if sheet.max_row < 1 or sheet.max_column < 1:
            continue
        
        # Check if sheet has headers
        headers = [cell.value for cell in sheet[1]]
        if all(h is None for h in headers):
            continue
            
        # Ensure we have minimum dimensions for table (at least header row)
        if sheet.max_row == 1 and sheet.max_column > 0:
            # Add empty row for table structure
            for col_idx in range(1, sheet.max_column + 1):
                sheet.cell(row=2, column=col_idx, value="")
        
        # Create table if sheet has data
        if sheet.max_column > 0:
            table_ref = f'A1:{get_column_letter(sheet.max_column)}{max(sheet.max_row, 2)}'
            tab = Table(
                displayName=f"Table_{sheet.title.replace(' ', '_').replace('-', '_')}",
                ref=table_ref
            )
            tab.tableStyleInfo = style
            sheet.add_table(tab)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set reasonable width limits
            adjusted_width = min(max(max_length + 2, 12), 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply text wrapping to analysis columns
        # Find analysis columns by checking header row
        if sheet.max_row >= 1:
            analysis_columns = []
            for col_idx in range(1, sheet.max_column + 1):
                header_value = str(sheet.cell(row=1, column=col_idx).value or '').lower()
                if header_value in ['analysis', 'hc_contractor_summary', 'cbx_contractor_summary']:
                    analysis_columns.append(col_idx)
            
            # Apply formatting to analysis columns
            for col_idx in analysis_columns:
                col_letter = get_column_letter(col_idx)
                
                # Set wider width for analysis columns
                sheet.column_dimensions[col_letter].width = 40
                
                # Apply text wrapping
                for row_idx in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(wrapText=True, vertical='top')
    
    # Save the formatted workbook
    wb.save(output_file)
    print(f"✅ Applied formatting: {len(sheets)} sheets, tables, filters, and styling")

def main():
    import sys
    
    if len(sys.argv) != 3:
        print("Usage: python3 format_excel.py <input_file> <output_file>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    apply_excel_formatting(input_file, output_file)

if __name__ == "__main__":
    main()
